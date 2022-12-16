
from openpyxl import Workbook, load_workbook

wb = Workbook()
ws = wb.active

colunas = ('A','B','C','D')
linhas = range(2,101)

for x in linhas:
    if x == 2:
        ws[f'D{x}'] = f'=C{x}'
    else:
        ws[f'D{x}'] = f'=C{x}+D{x-1}'

wb.save('extrato-bpi-VAZIA.xlsx')
print('Planilha atualizada com sucesso')
