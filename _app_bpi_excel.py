import json
import time
from pathlib import Path
from operator import itemgetter
from datetime import date, datetime
from openpyxl import Workbook, load_workbook


# Leitura do arquivo JSON gerado pelo _auto-bpi > descarrega em lista_extrato
# with open('extrato-bpi.json') as extrato_json:
#     lista_extrato = json.loads(extrato_json)
# extrato_json.close()

lista_extrato = json.load(x:=open("extrato-bpi.json"))
tam = len(lista_extrato)

# Abrir o arquivo Excel e ativar a planilha 1
arq_excel = load_workbook(filename="extrato-bpi.xlsx")
plan1 = arq_excel.active

# Encontrar a primeira linha vazia na planilha
linini = plan1.max_row + 1
print('A primeira linha vazia é ', linini)
print('Formatando os dados de extrato para atualização da planilha...')
time.sleep(1)

#  transforma data em objeto datetime e transforma valor e saldo string para float
for i, conteudo in enumerate(lista_extrato):
    try:
        conteudo = datetime.strptime(conteudo, '%d-%m-%Y')
        lista_extrato[i] = date(conteudo.year, conteudo.month, conteudo.day)

    except ValueError:
        if conteudo.endswith('EUR'):
            conteudo = conteudo.rstrip(' EUR')
            conteudo = conteudo.replace('.', '')
            conteudo = conteudo.replace(',', '.')
            lista_extrato[i] = float(conteudo)
del i, conteudo
print('Elementos do extrato transformados para o tipo de dado correto')
time.sleep(1)

# Elimina datas duplas passadas pelo extrato BPI. O objeto é tipo datetime.date
# Apenas temos que colocar a segunda parte "date" na comparação do isinstance, 
# senão não funciona

for i in range(tam):
    try:
        if isinstance(lista_extrato[i], date) and isinstance(lista_extrato[i+1], date):
            del lista_extrato[i+1]
    except IndexError:
        break
print('Datas duplicadas eliminadas')
time.sleep(1)

# Faz a lista de listas com lista_extrato dividida em linhas de 4 colunas
listnew_extrato = [['']]
try:
    for i in range(0,tam,4):
        listnew_extrato.append([ lista_extrato[i], lista_extrato[i+1], 
        lista_extrato[i+2], lista_extrato[i+3]])
except:
    pass
finally:
    # Apaga indice zero provisório e despeja listnew em lista_extrato
    listnew_extrato.pop(0)
    lista_extrato = listnew_extrato
    del listnew_extrato
    tam = len(lista_extrato)        #<<<atualiza tam - tamanho da lista extrato
print('Lista extrato organizada em linhas e colunas')    
time.sleep(1)

# Apaga a quarta coluna de saldo
for i in range(tam):
    del lista_extrato[i][3]

# Ordena a lista extrato por data crescente (indice 0 das listas)
lista_new = sorted(lista_extrato, key=lambda x: x[0])
lista_extrato = lista_new
del lista_new
print('Apaga os saldos e ordena linhas por data')
time.sleep(1)

# Checa a planilha para encontrar linhas repetidas
# Compara o campo descrição de lista_extrato com a planilha.

# for i in range(tam):
#     for lin in plan1.iter_rows():
#         if lista_extrato[i][0] == plan1.cell[f'A{lin}'].value:
#             if lista_extrato[i][1] == plan1.cell[f'B{lin}'].value:
#                 del lista_extrato[i]


# Copia da lista extrato para a planilha
i = 0
linfin = linini + tam
for lin in range(linini, linfin):
    try:
        for col in range(1,4):
            plan1.cell(row=lin, column=col, value=lista_extrato[i][col-1])
    except IndexError:
        break
    i += 1
    
pathxl = 'extrato-bpi.xlsx'
arq_excel.save(pathxl)
print(f'Planilha {pathxl} atualizada da linha {linini} até a {linfin}')

opcao = input('Deseja apagar o arquivo intermediario JSON? ')
if opcao == "S" or opcao == "s":
    extrato_json = Path('extrato-bpi.json')
    extrato_json.unlink()
